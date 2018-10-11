# new_member_script.py
from openpyxl import load_workbook, Workbook
import csv

def parse_invitees(file_name):
    wb = load_workbook(file_name)
    sheet = wb.worksheets[0]

    invitees = []
    for i, invitee_row in enumerate(sheet):
        if i >= 1:
            invitees.append((
                invitee_row[0].value, # first name
                invitee_row[1].value, # last name
                invitee_row[2].value, # year
                invitee_row[3].value, # email
            ))

    return invitees

def parse_member_emails(file_name):
    wb = load_workbook(file_name)
    sheet = wb.worksheets[0]

    members = []
    for i, member_row in enumerate(sheet):
        if i >= 1:
            members.append(member_row[0].value)

    return members

def write_invite_list(invitees):
    with open('to_invite.csv', mode='w') as invite_file:
        writer = csv.writer(
            invite_file,
            delimiter=',',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL)

        writer.writerow(['First Name', 'Last Name', 'Year', 'Email'])
        for invitee in invitees:
            writer.writerow(invitee)

def get_to_invite_list(invitees, member_emails):
    to_invite = []

    for invitee in invitees:
        email = invitee[3] # get email from tuple

        if(not (email in member_emails)):
            to_invite.append(invitee)

    return to_invite

# parse out workbooks
print("parsing invitees.xlsx...")
invitees = parse_invitees('invitees.xlsx')
print("done parsing invitees.xlsx")

print("parsing members.xlsx...")
member_emails = parse_member_emails('members.xlsx')
print("done parsing members.xlsx")

print("removing existing members...")
to_invite = get_to_invite_list(invitees, member_emails) # parse out collisions
print("number of new members to invite: " + len(to_invite).__str__())

print("writing members to invite to to_invite.csv")
write_invite_list(to_invite)                # write list to 'to_invite.csv'
print("task completed!")
